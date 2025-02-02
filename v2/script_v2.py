"""Schedule Processor: iCal/Excel generator."""

import csv
import datetime
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Iterable, Iterator, List, Optional, Sequence

from zoneinfo import ZoneInfo

import ics
import openpyxl
import rich
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ==================== Configuration ====================
class ScheduleConfig:
    """Central configuration container for schedule processing."""

    # Processing flags
    ADD_BREAKS: bool = False
    ADD_MARK: bool = True

    # Date configuration
    FIRST_DATE: datetime.date = datetime.date(2025, 2, 3)  # Monday
    TIMEZONE: ZoneInfo = ZoneInfo("Europe/Moscow")

    # File structure
    EXPECTED_FILENAME_PARTS: Sequence[int] = (4, 5)
    SCHEDULE_COLUMN_INDEX: int = 3
    VALID_TYPES: Sequence[str] = ("л", "с")
    # Для соответствия ключам в RINGS
    TYPE_MAP: dict[str, str] = {"л": "l", "с": "s"}

    # Time slots configuration
    # Каждое значение списка преобразуем в словарь с ключами:
    # "start", "end", "break_start", "break_end", "numbers"
    RINGS: dict[str, dict[str, List[Any]]] = {
        "s": {
            "9:00": [
                datetime.time(9, 0),
                datetime.time(10, 30),
                datetime.time(10, 45),
                datetime.time(12, 15),
                "1,2",
            ],
            "13:10": [
                datetime.time(13, 10),
                datetime.time(14, 40),
                datetime.time(14, 55),
                datetime.time(16, 25),
                "3,4",
            ],
        },
        "l": {
            "9:00": [
                datetime.time(9, 0),
                datetime.time(9, 45),
                datetime.time(9, 50),
                datetime.time(10, 35),
                "1",
            ],
            "10:55": [
                datetime.time(10, 55),
                datetime.time(11, 40),
                datetime.time(11, 45),
                datetime.time(12, 30),
                "2",
            ],
            "13:10": [
                datetime.time(13, 10),
                datetime.time(13, 55),
                datetime.time(14, 0),
                datetime.time(14, 45),
                "3",
            ],
            "15:00": [
                datetime.time(15, 0),
                datetime.time(15, 45),
                datetime.time(15, 50),
                datetime.time(16, 35),
                "4",
            ],
            "16:45": [
                datetime.time(16, 45),
                datetime.time(17, 30),
                datetime.time(17, 35),
                datetime.time(18, 20),
                "5",
            ],
        },
    }

    # Weekday mappings
    WEEK_DAYS: dict[str, int] = {
        "пн": 0,
        "вт": 1,
        "ср": 2,
        "чт": 3,
        "пт": 4,
        "сб": 5,
        "вс": 6,
    }
    WEEK_DAYS_INVERTED: dict[int, str] = {
        0: "Пн",
        1: "Вт",
        2: "Ср",
        3: "Чт",
        4: "Пт",
        5: "Сб",
        6: "Вс",
    }

    # Excel styling configuration
    WIDTH_COLUMNS: List[int] = [8, 12, 4, 4, 16, 4, 20]

    STYLES: dict[str, dict[str, Any]] = {
        "header": {
            "font": Font(size=14, name="Roboto"),
            "fill": PatternFill("solid", start_color="D9EAD3"),
            "border": Border(
                top=Side(style="thin"),
                bottom=Side(style="thick"),
                left=Side(style="thin"),
                right=Side(style="thin"),
            ),
            "alignment": Alignment(horizontal="center"),
        },
        "lecture": {
            "fill": PatternFill("solid", start_color="FFE6E6"),
            "font": Font(name="Roboto"),
        },
        "seminar": {
            "fill": PatternFill("solid", start_color="E6FFE6"),
            "font": Font(name="Roboto"),
        },
        "week_number": {"font": Font(bold=True, size=30, name="Roboto")},
        "day_name": {"font": Font(bold=True, name="Roboto")},
    }

    # iCal settings
    LOCATION: str = "Piskarovskiy Ave, 47, Sankt-Peterburg, Russia, 195067"
    CATEGORIES: dict[str, str] = {"л": "Лекция", "с": "Семинар"}

    # Output configuration
    OUTPUT_DIR: Path = Path("output")
    INPUT_DIR: Path = Path("input")

    @classmethod
    def validate_config(cls) -> None:
        """Validate configuration values."""
        if not cls.INPUT_DIR.exists():
            raise FileNotFoundError(f"Input directory not found: {cls.INPUT_DIR}")

        if not cls.OUTPUT_DIR.exists():
            cls.OUTPUT_DIR.mkdir(parents=True)

        if cls.FIRST_DATE.weekday() != 0:
            raise ValueError("FIRST_DATE must be a Monday")

        for rings in cls.RINGS.values():
            for times in rings.values():
                if len(times) != 5:
                    raise ValueError("Each time slot must have exactly 5 elements")
                if not isinstance(times[4], str):
                    raise TypeError("Last element in time slot must be a string")


# ==================== Data Models ====================
class ScheduleEvent:
    """Data class representing a single schedule event."""

    __slots__ = ["week", "date", "time_slot", "numbers", "type", "subject"]

    def __init__(
        self,
        week: int,
        date: datetime.date,
        time_slot: dict[str, Any],
        event_type: str,
        subject: str,
    ) -> None:
        self.week = week
        self.date = date
        self.time_slot = time_slot  # Ожидается словарь с ключами "start", "end", "break_start", "break_end", "numbers"
        self.numbers = time_slot["numbers"]
        self.type = event_type
        self.subject = subject


# ==================== File Processing ====================
class FileProcessor(ABC):
    """Abstract base class for file processors."""

    def __init__(self, config: ScheduleConfig) -> None:
        self.config = config
        self.weekdays = config.WEEK_DAYS

    @abstractmethod
    def process(self, file_path: Path) -> List[List[Any]]:
        """Process file and return raw schedule data."""
        pass

    def _process_row(self, row: List[Any], previous_row: List[Any]) -> List[Any]:
        """Common row processing logic."""
        processed_row: List[Any] = []
        for idx, value in enumerate(row):
            if not value and previous_row:
                processed_row.append(previous_row[idx])
            else:
                processed_row.append(self._process_cell(value, idx))
        return processed_row

    def _process_cell(self, value: Any, column_idx: int) -> Any:
        """Process individual cell based on column index."""
        if column_idx == 0 and value:
            return self.weekdays.get(str(value).lower().strip(), value)
        if column_idx == self.config.SCHEDULE_COLUMN_INDEX and value:
            return self._parse_schedule_cell(value)
        return value

    def _parse_schedule_cell(self, cell: str) -> List[int]:
        """Parse schedule cell with interval expansion."""
        def expand(interval: str) -> List[int]:
            start, end = map(int, interval.split("-"))
            return list(range(start, end + 1))

        return [
            num
            for part in cell.replace(" ", "").split(",")
            for num in (expand(part) if "-" in part else [int(part)])
        ]


class CSVProcessor(FileProcessor):
    """CSV file processor implementation."""

    def process(self, file_path: Path) -> List[List[Any]]:
        with file_path.open(encoding="utf-8") as f:
            reader = csv.reader(f)
            return self._process_file(reader)

    def _process_file(self, reader: Iterable[List[str]]) -> List[List[Any]]:
        data: List[List[Any]] = []
        prev_row = next(reader)  # Считаем, что первая строка — заголовок
        for row in reader:
            processed = self._process_row(row, prev_row)
            data.append(processed)
            prev_row = processed.copy()
        return data


class XLSXProcessor(FileProcessor):
    """XLSX file processor implementation."""

    def process(self, file_path: Path) -> List[List[Any]]:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        rows = list(self._iter_sheet_rows(sheet))
        if not rows:
            return []
        data: List[List[Any]] = []
        prev_row = rows[0]
        for row in rows[1:]:
            processed = self._process_row(row, prev_row)
            data.append(processed)
            prev_row = processed.copy()
        return data

    def _iter_sheet_rows(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> Iterator[List[Any]]:
        """Iterate over sheet rows with proper start position."""
        start_row = 2 if sheet.cell(row=1, column=1).value == "пн" else 1
        return (
            [cell.value for cell in row]
            for row in sheet.iter_rows(min_row=start_row)
            if any(cell.value for cell in row)
        )


# ==================== Schedule Generation ====================
class ScheduleGenerator:
    """Main schedule processing and coordination class."""

    def __init__(self, config: ScheduleConfig) -> None:
        self.config = config
        self.processors = {
            ".csv": CSVProcessor(config),
            ".xlsx": XLSXProcessor(config),
        }
        # Структура данных: тип события -> identifier -> List[ScheduleEvent]
        self.data: dict[str, dict[tuple, List[ScheduleEvent]]] = {"л": {}, "с": {}}

    def process_files(self, input_dir: Optional[Path] = None) -> None:
        """Process all files in input directory."""
        input_dir = input_dir or self.config.INPUT_DIR
        for ext, processor in self.processors.items():
            for file_path in input_dir.glob(f"*{ext}"):
                self._process_single_file(file_path, processor)

    def _process_single_file(self, file_path: Path, processor: FileProcessor) -> None:
        """Process individual file and store results."""
        filename = file_path.stem.lower()
        parts = filename.split("_")
        if len(parts) not in self.config.EXPECTED_FILENAME_PARTS:
            raise ValueError(f"Invalid filename format: {filename}")
        event_type = parts[0]
        if event_type not in self.config.VALID_TYPES:
            raise ValueError(f"Unknown schedule type: {event_type}")

        identifier = self._create_identifier(parts, event_type)
        schedule_data = processor.process(file_path)
        events = self._transform_data(schedule_data, event_type)
        self.data[event_type][identifier] = events

        rich.print(f"[green]Processed:[/green] {file_path.name}")

    def _create_identifier(self, parts: List[str], event_type: str) -> tuple:
        """Create unique identifier for schedule entry."""
        base_id = tuple(parts[1:3])
        return base_id if event_type == "л" else (*base_id, parts[4])

    def _transform_data(self, raw_data: List[List[Any]], event_type: str) -> List[ScheduleEvent]:
        """Convert raw data to ScheduleEvent objects."""
        # Определяем максимальную неделю из данных
        max_week = max(max(row[self.config.SCHEDULE_COLUMN_INDEX]) for row in raw_data)
        current_date = self.config.FIRST_DATE
        week = 1
        events: List[ScheduleEvent] = []

        # Для доступа к "Рингам" используем маппинг типов
        ring_type = self.config.TYPE_MAP[event_type]
        while week <= max_week:
            for row in raw_data:
                if week in row[self.config.SCHEDULE_COLUMN_INDEX]:
                    # Получаем временной интервал по ключу времени (предполагается, что row[1] содержит время начала в виде строки)
                    time_slot_raw = self.config.RINGS[ring_type].get(row[1])
                    if not time_slot_raw:
                        raise ValueError(f"Invalid time key {row[1]} for event type {event_type}")
                    # Преобразуем список в словарь
                    time_slot = {
                        "start": time_slot_raw[0],
                        "end": time_slot_raw[1],
                        "break_start": time_slot_raw[2],
                        "break_end": time_slot_raw[3],
                        "numbers": time_slot_raw[4],
                    }
                    events.append(
                        ScheduleEvent(
                            week=week,
                            date=current_date,
                            time_slot=time_slot,
                            event_type=event_type,
                            subject=row[2],
                        )
                    )
            current_date += datetime.timedelta(days=1)
            # Если прошла неделя (после воскресенья), увеличиваем номер недели
            if current_date.weekday() == 0:
                week += 1
        return sorted(events, key=lambda e: (e.date, e.time_slot["start"]))


# ==================== Export Modules ====================
class ExcelExporter:
    """Excel export handler with advanced styling."""

    def __init__(self, config: ScheduleConfig) -> None:
        self.config = config
        self.styles = self.config.STYLES

    def export(self, events: List[ScheduleEvent], identifier: str) -> None:
        """Generate Excel file from schedule events."""
        wb = openpyxl.Workbook()
        ws = self._prepare_worksheet(wb.active, identifier)
        self._add_data(ws, events)
        self._apply_final_adjustments(ws)
        self._save_workbook(wb, identifier)

    def _prepare_worksheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, identifier: str) -> openpyxl.worksheet.worksheet.Worksheet:
        """Set up worksheet headers and metadata."""
        ws.title = "Schedule"
        ws.append([f"{identifier} | Generated by Schedule Processor"])
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(size=14, name="Roboto", bold=True)
        return ws

    def _add_data(self, ws: openpyxl.worksheet.worksheet.Worksheet, events: List[ScheduleEvent]) -> None:
        """Add schedule data to worksheet with proper formatting."""
        ws.append(["Week", "Date", "Day", "Numbers", "Time", "Type", "Subject"])
        self._style_header(ws[2])
        for event in events:
            row = [
                event.week,
                event.date.isoformat(),
                event.date.strftime("%a"),
                ",".join(map(str, event.numbers)),
                self._format_time_range(event),
                "Л" if event.type == "л" else "С",
                event.subject,
            ]
            ws.append(row)
            self._style_data_row(ws[ws.max_row], event.type)

    def _format_time_range(self, event: ScheduleEvent) -> str:
        """Format time range for Excel display."""
        start_dt = datetime.datetime.combine(event.date, event.time_slot["start"])
        end_dt = datetime.datetime.combine(event.date, event.time_slot["end"])
        return " - ".join([start_dt.strftime("%H:%M"), end_dt.strftime("%H:%M")])

    def _style_header(self, header_row: Sequence[openpyxl.cell.cell.Cell]) -> None:
        """Apply styles to header row."""
        for cell in header_row:
            cell.font = self.styles["header"]["font"]
            cell.fill = self.styles["header"]["fill"]
            cell.border = self.styles["header"]["border"]
            cell.alignment = Alignment(horizontal="center")

    def _style_data_row(self, row: Sequence[openpyxl.cell.cell.Cell], event_type: str) -> None:
        """Apply styles to data row based on event type."""
        style = self.styles["lecture"] if event_type == "л" else self.styles["seminar"]
        # Применяем заливку к колонкам Time, Type, Subject (индексы 4, 5, 6)
        for cell in row[4:7]:
            cell.fill = style["fill"]
        # Особое оформление для номера недели
        row[0].font = Font(bold=True, size=30, name="Roboto")
        # Выравнивание по центру для всех ячеек, кроме Subject
        for cell in row[:-1]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _apply_final_adjustments(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Apply final formatting and adjustments to worksheet."""
        # Устанавливаем ширину колонок
        for i, width in enumerate(self.config.WIDTH_COLUMNS):
            ws.column_dimensions[chr(65 + i)].width = width
        self._merge_date_cells(ws)
        self._merge_week_cells(ws)

    def _merge_date_cells(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Merge cells for same dates."""
        merge_start = 3
        prev_date = ws[f"B{merge_start}"].value
        for row_idx in range(merge_start + 1, ws.max_row + 1):
            current_date = ws[f"B{row_idx}"].value
            if current_date != prev_date:
                if merge_start < row_idx - 1:
                    ws.merge_cells(f"B{merge_start}:B{row_idx - 1}")
                    ws.merge_cells(f"C{merge_start}:C{row_idx - 1}")
                merge_start = row_idx
                prev_date = current_date
        if merge_start < ws.max_row:
            ws.merge_cells(f"B{merge_start}:B{ws.max_row}")
            ws.merge_cells(f"C{merge_start}:C{ws.max_row}")

    def _merge_week_cells(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Merge cells for same weeks."""
        merge_start = 3
        prev_week = ws[f"A{merge_start}"].value
        for row_idx in range(merge_start + 1, ws.max_row + 1):
            current_week = ws[f"A{row_idx}"].value
            if current_week != prev_week:
                if merge_start < row_idx - 1:
                    ws.merge_cells(f"A{merge_start}:A{row_idx - 1}")
                merge_start = row_idx
                prev_week = current_week
        if merge_start < ws.max_row:
            ws.merge_cells(f"A{merge_start}:A{ws.max_row}")

    def _save_workbook(self, wb: openpyxl.workbook.workbook.Workbook, identifier: str) -> None:
        """Save workbook to file."""
        output_dir = self.config.OUTPUT_DIR
        output_dir.mkdir(exist_ok=True)
        filename = output_dir / f"{identifier}.xlsx"
        wb.save(filename)
        rich.print(f"[green]Saved Excel file:[/green] {filename}")


class ICalExporter:
    """iCalendar export handler."""

    def __init__(self, config: ScheduleConfig) -> None:
        self.config = config

    def export(self, events: List[ScheduleEvent], identifier: str) -> None:
        """Generate iCal file from schedule events."""
        calendar = ics.Calendar()
        for event in events:
            for ical_event in self._create_events(event):
                calendar.events.add(ical_event)
        self._save_calendar(calendar, identifier)

    def _create_events(self, event: ScheduleEvent) -> set[ics.Event]:
        """Create iCal events from schedule event."""
        events_set: set[ics.Event] = set()
        base_event = self._create_base_event(event)
        events_set.add(base_event)
        if self.config.ADD_BREAKS:
            # Используем break_start и break_end
            events_set.add(
                self._create_break_event(
                    event, event.time_slot["break_start"], event.time_slot["break_end"]
                )
            )
        return events_set

    def _create_base_event(self, event: ScheduleEvent) -> ics.Event:
        """Create main event for the schedule entry."""
        event_type = "Л" if event.type == "л" else "С"
        event_name = f"{event_type} {event.subject}"
        ical_event = ics.Event(
            name=event_name,
            begin=datetime.datetime.combine(
                event.date, event.time_slot["start"], tzinfo=self.config.TIMEZONE
            ),
            end=datetime.datetime.combine(
                event.date, event.time_slot["end"], tzinfo=self.config.TIMEZONE
            ),
            location=self.config.LOCATION,
            categories=[self.config.CATEGORIES[event.type], "СЗГМУ"],
            created=datetime.datetime.now(tz=self.config.TIMEZONE),
        )
        if self.config.ADD_MARK:
            ical_event.description = "https://t.me/mechnews_1k"
        return ical_event

    def _create_break_event(
        self, event: ScheduleEvent, start: datetime.time, end: datetime.time
    ) -> ics.Event:
        """Create break event between sessions."""
        return ics.Event(
            name="Перерыв",
            begin=datetime.datetime.combine(event.date, start, tzinfo=self.config.TIMEZONE),
            end=datetime.datetime.combine(event.date, end, tzinfo=self.config.TIMEZONE),
            location=self.config.LOCATION,
            categories=["Перерыв", "СЗГМУ"],
            transparent=False,
        )

    def _save_calendar(self, calendar: ics.Calendar, identifier: str) -> None:
        """Save iCal file to disk."""
        output_dir = self.config.OUTPUT_DIR
        output_dir.mkdir(exist_ok=True)
        filename = output_dir / f"{identifier}.ics"
        with filename.open("wb") as f:
            f.write(calendar.serialize().encode("utf-8"))
        rich.print(f"[green]Saved iCal file:[/green] {filename}")


class ExportManager:
    """Manages export to different formats."""

    def __init__(self, config: ScheduleConfig) -> None:
        self.config = config
        self.exporters = {
            "excel": ExcelExporter(config),
            "ical": ICalExporter(config),
        }

    def export_all(self, schedule: List[ScheduleEvent], identifier: str) -> None:
        """Export schedule to all available formats."""
        for name, exporter in self.exporters.items():
            try:
                exporter.export(schedule, identifier)
            except Exception as e:
                rich.print(f"[red]Error exporting {name}:[/red] {str(e)}")


# ==================== Main Execution ====================
def main() -> None:
    """Main processing pipeline."""
    ScheduleConfig.validate_config()
    config = ScheduleConfig()
    processor = ScheduleGenerator(config)
    processor.process_files()
    exporter = ExportManager(config)
    # Предполагается, что для идентификатора расписание для семинаров и лекций имеют одинаковый base_id
    for identifier, seminar_events in processor.data["с"].items():
        lecture_events = processor.data["л"].get(identifier[:2], [])
        combined = lecture_events + seminar_events
        exporter.export_all(combined, "_".join(identifier).upper())


if __name__ == "__main__":
    main()
