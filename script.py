"""Script for processing CSV files with schedule data."""

import csv
import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import ics
import openpyxl
import rich

ADD_BREAKS: bool = False
ADD_MARK: bool = True
FIRST_DATE = datetime.date(2025, 2, 3)  # Monday
EXPECTED_FILENAME_PARTS: tuple[int] = (4, 5)
SCHEDULE_COLUMN_INDEX = 3
RINGS = {
    "s": {
        "9:00": [datetime.time(9, 0), datetime.time(10, 30), datetime.time(10, 45), datetime.time(12, 15), '1,2'],
        "13:10": [datetime.time(13, 10), datetime.time(14, 40), datetime.time(14, 55), datetime.time(16, 25),'3,4'],
    },
    "l": {
        "9:00": [datetime.time(9, 0), datetime.time(9, 45), datetime.time(9, 50), datetime.time(10, 35),'1'],
        "10:55": [datetime.time(10, 55), datetime.time(11, 40), datetime.time(11, 45), datetime.time(12, 30),'2'],
        "13:10": [datetime.time(13, 10), datetime.time(13, 55), datetime.time(14, 0), datetime.time(14, 45),'3'],
        "15:00": [datetime.time(15, 0), datetime.time(15, 45), datetime.time(15, 50), datetime.time(16, 35),'4'],
        "16:45": [datetime.time(16, 45), datetime.time(17, 30), datetime.time(17, 35), datetime.time(18, 20),'5'],
    },
}
WIDTH_COLUMNS = [8, 12, 4, 4, 16, 4, 20]

WEEK_DAYS = {'пн': 0, 'вт': 1, 'ср': 2, 'чт': 3, 'пт': 4, 'сб': 5, 'вс': 6}
WEEK_DAYS_INVERTED = {
    0: 'Пн',
    1: 'Вт',
    2: 'Ср',
    3: 'Чт',
    4: 'Пт',
    5: 'Сб',
    6: 'Вс'
}


def main() -> None:
    """Execute the main schedule processing functionality."""
    l_schedules: dict[str, list] = {}
    s_schedules: dict[str, list] = {}

    # Process CSV files
    for file_path in Path("./input").glob("*.csv"):
        file_name: str = file_path.stem.lower()
        filename_parts = file_name.split("_")
        id_ = tuple(filename_parts[1:3])

        if len(filename_parts) in EXPECTED_FILENAME_PARTS:
            schedule_data = process_csv_file(file_path)
            if filename_parts[0] == "л":
                l_schedules[id_] = schedule_data
            elif filename_parts[0] == "с":
                id_ += (filename_parts[4],)
                s_schedules[id_] = schedule_data
            else:
                error_msg = f"Unknown file name: {file_name}"
                raise ValueError(error_msg)
        rich.print(f"Processed file: {file_path}")

    # Process XLSX files
    for file_path in Path("./input").glob("*.xlsx"):
        file_name: str = file_path.stem.lower()
        filename_parts = file_name.split("_")
        id_ = tuple(filename_parts[1:3])

        if len(filename_parts) in EXPECTED_FILENAME_PARTS:
            schedule_data = process_xlsx_file(file_path)
            if filename_parts[0] == "л":
                l_schedules[id_] = schedule_data
            elif filename_parts[0] == "с":
                id_ += (filename_parts[4],)
                s_schedules[id_] = schedule_data
            else:
                error_msg = f"Unknown file name: {file_name}"
                raise ValueError(error_msg)
        rich.print(f"Processed file: {file_path}")

    # Generate right schedule
    for s_id, s_schedule in s_schedules.items():
        n_s_sch = gen_schedule(s_schedule, "s")
        n_l_sch = gen_schedule(l_schedules[s_id[:2]], "l")
        n_sch = n_s_sch + n_l_sch
        gen_excel_file(n_sch, s_id)
        gen_ical(n_sch, s_id)


def expand_interval(interval: str) -> list[int]:
    """Expand a range interval into a list of integers.

    Args:
        interval (str): The interval string in format "start-end"

    Returns:
        list[int]: List of integers in the interval
    """
    start, end = map(int, interval.split("-"))
    return list(range(start, end + 1))


def process_schedule_cell(cell: str) -> list[int]:
    """Process a schedule cell and expand any intervals.

    Args:
        cell (str): The schedule cell content

    Returns:
        list: Processed schedule data with expanded intervals
    """
    values = cell.replace(" ", "").split(",")
    result: list[int] = []

    for value in values:
        if "-" in value:
            result.extend(expand_interval(value))
        else:
            result.append(int(value) if value.isdigit() else value)

    return result


def process_csv_file(file_path: Path) -> list:
    """Process a single CSV file and return the schedule data.

    Args:
        file_path (Path): The path to the CSV file to process.

    Returns:
        list: The processed schedule data, excluding the header row.
    """
    schedule_data: list[list[str | int | list[int]]] = []

    with file_path.open() as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=",")
        previous_row = next(csv_reader)  # Read header
        schedule_data.append(previous_row)

        for current_row in csv_reader:
            for index, cell in enumerate(current_row):
                if not cell and previous_row:
                    current_row[index] = previous_row[index]
                if index == SCHEDULE_COLUMN_INDEX and cell:
                    current_row[index] = process_schedule_cell(cell)
                elif index == 0 and cell:
                    current_row[index] = WEEK_DAYS[cell.lower()]

            previous_row = current_row.copy()
            schedule_data.append(current_row)

    return schedule_data[1:]


def gen_schedule(sch: list[list[str | int | list[int]]], type_: str) -> list[list[str | int]]:
    """Generate a schedule with expanded time slots and dates.

    Args:
        sch: Raw schedule data
        type_: Schedule type identifier

    Returns:
        Processed schedule with dates and times
    """
    from operator import itemgetter

    MAX_WEEKDAY = 6
    processed_schedule = []
    max_week = max(max(row[SCHEDULE_COLUMN_INDEX]) for row in sch)
    current_day = FIRST_DATE
    week = 1

    while week <= max_week:
        for row in sch:
            if week in row[SCHEDULE_COLUMN_INDEX] and current_day.weekday() == row[0]:
                time_slot = RINGS[type_][row[1]]
                processed_schedule.append([
                    week,
                    current_day,
                    time_slot,
                    RINGS[type_][row[1]][4],
                    type_,
                    row[2],
                ])

        current_day += datetime.timedelta(days=1)
        if current_day.weekday() == MAX_WEEKDAY:
            week += 1

    processed_schedule.sort(key=itemgetter(1, 2))
    return processed_schedule


def prepocess_schedule_for_excel(schedule_data: list[list[str | int | list[int]]]) -> list[list[str | int | list[int]]]:
    schedule_data = schedule_data.copy()
    """Prepare schedule data for Excel export."""
    for i in range(len(schedule_data)):
        row = [
            schedule_data[i][0],
            schedule_data[i][1].isoformat(),
            WEEK_DAYS_INVERTED[schedule_data[i][1].weekday()],
            schedule_data[i][3],
            " - ".join([schedule_data[i][2][0].isoformat(timespec='minutes'), schedule_data[i][2][3].isoformat(timespec='minutes')]),
            'Л' if schedule_data[i][4] == 'l' else 'С',
            schedule_data[i][5]
        ]
        schedule_data[i] = row
    schedule_data.sort()
    return schedule_data


def gen_excel_file(schedule_data: list[list[str | int | list[int]]], id_: list) -> None:
    """Generate an Excel file from the schedule data."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "made by Мечечные новости ;)"

    # Add title row
    id_: str = "_".join(id_).upper()
    worksheet.append([f"{id_}. MechNews: https://t.me/mechnews_1k"])
    worksheet.merge_cells("A1:G1")
    worksheet["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")
    worksheet["A1"].font = openpyxl.styles.Font(size=14, name="Roboto")

    # Add header
    header = ["нед.", "Дата", "Д/н", "№", "Время", "Тип", "Предмет"]
    worksheet.append(header)
    for cell in worksheet[2]:
        cell.font = openpyxl.styles.Font(size=14, name="Roboto")
        cell.fill = openpyxl.styles.PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        cell.border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style="thin"),
            left=openpyxl.styles.Side(style="thin"),
            right=openpyxl.styles.Side(style="thin"),
            bottom=openpyxl.styles.Side(style="thick")
        )

    schedule_data = prepocess_schedule_for_excel(schedule_data)

    # Merge cells for the first row
    prev_date = None
    merge_start_date = 3
    merge_start_week = 3
    current_row = 3
    prev_week = None

    # define styles
    thin_border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style="thin"))
    thick_border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style="thick"))
    lecture_fill = openpyxl.styles.PatternFill(
        start_color="FFE6E6",
        end_color="FFE6E6",
        fill_type="solid"
        )
    seminar_fill = openpyxl.styles.PatternFill(
        start_color="E6FFE6",
        end_color="E6FFE6",
        fill_type="solid"
        )

    # Main loop
    for row in schedule_data:
        worksheet.append(row)
        current_cells = worksheet[current_row]
        # Base styling
        for i in range(len(current_cells)):
            # Font
            if i == 0:

                current_cells[i].font = openpyxl.styles.Font(bold=True, size=30, name="Roboto")
            elif i == 2:

                current_cells[i].font = openpyxl.styles.Font(bold=True, name="Roboto")
            else:

                current_cells[i].font = openpyxl.styles.Font(name="Roboto")
            # Alignment
            if i == 6:
                pass
            else:

                current_cells[i].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        # Apply colors based on type
        if row[5] == "Л":
            current_cells[5].fill = lecture_fill
            current_cells[4].fill = lecture_fill
            current_cells[6].fill = lecture_fill
        else:
            current_cells[5].fill = seminar_fill
            current_cells[4].fill = seminar_fill
            current_cells[6].fill = seminar_fill

        # Apply borders and merging based on date and week
        # Date
        if prev_date and prev_date != row[1]:
            if merge_start_date < current_row - 1:
                worksheet.merge_cells(f"B{merge_start_date}:B{current_row - 1}")
                worksheet.merge_cells(f"C{merge_start_date}:C{current_row - 1}")
            merge_start_date = current_row
            # Apply thick border between days
            for col in range(1, 8):
                cell = worksheet.cell(row=current_row-1, column=col)
                cell.border = thin_border
        # Week
        if prev_week and prev_week != row[0]:
            if merge_start_week < current_row - 1:

                worksheet.merge_cells(f"A{merge_start_week}:A{current_row - 1}")
                # Apply thick border between weeks
                for col in range(1, 8):
                    cell = worksheet.cell(row=current_row - 1, column=col)
                    cell.border = thick_border
            merge_start_week = current_row
        prev_date = row[1]
        prev_week = row[0]
        current_row += 1

    # Merge the last group if needed
    if merge_start_date < current_row - 1:
        worksheet.merge_cells(f"B{merge_start_date}:B{current_row - 1}")
        worksheet.merge_cells(f"C{merge_start_date}:C{current_row - 1}")
    if merge_start_week < current_row - 1:
        worksheet.merge_cells(f"A{merge_start_week}:A{current_row - 1}")
    for i, width in enumerate(WIDTH_COLUMNS):
        column_letter = chr(65 + i)
        worksheet.column_dimensions[column_letter].width = width

    # Save the workbook
    filename = f"output/{id_}.xlsx"
    rich.print(f"Saving file: {filename}")
    workbook.save(filename)


def gen_ical(schedule_data: list[list[str | int | list[int]]], id_: str) -> None:
    """Generate an iCal file from the schedule data.

    Args:
        schedule_data: Processed schedule data
        id_: Schedule identifier

    """
    id_: str = "_".join(id_).upper()
    schedule_data.sort()
    calendar: ics.Calendar = ics.Calendar()
    for row in schedule_data:
        date: datetime.date = row[1]
        time: list = row[2]
        type_: str = "Л" if row[4] == "l" else "С"
        name: str = row[5]
        data: str = ""  # str(row)
        for i in [0, 2]:
            event: ics.Event = ics.Event()
            if ADD_BREAKS:
                event.begin = datetime.datetime.combine(date, time[0 + i]).replace(tzinfo=ZoneInfo('Europe/Moscow'))
                event.end = datetime.datetime.combine(date, time[1 + i]).replace(tzinfo=ZoneInfo('Europe/Moscow'))
            elif i == 0:
                event.begin = datetime.datetime.combine(date, time[0]).replace(tzinfo=ZoneInfo('Europe/Moscow'))
                event.end = datetime.datetime.combine(date, time[3]).replace(tzinfo=ZoneInfo('Europe/Moscow'))
            else:
                continue
            event.name = f"{type_} {name}"
            if ADD_MARK:
                event.description = f"{data}\nhttps://t.me/mechnews_1k"
            else:
                event.description = f"{data}"
            event.location = "Piskarovskiy Ave, 47, Sankt-Peterburg, Russia, 195067"
            event.created = datetime.datetime.now(tz=ZoneInfo('Europe/Moscow'))
            calendar.events.add(event)
    filename: Path = Path(f"output/{id_}.ics")
    rich.print(f"Saving file: {filename} ({len(calendar.events)} events)")
    with filename.open("wb") as ical_file:
        ical_file.write(str(calendar.serialize()).encode('utf-8'))


if __name__ == "__main__":
    main()
