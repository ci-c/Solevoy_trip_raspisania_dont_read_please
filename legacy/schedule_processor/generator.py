"""Schedule generation and export functionality."""

import datetime
import logging
from pathlib import Path
from typing import List
from zoneinfo import ZoneInfo
import openpyxl
import ics
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from .config import RINGS, WEEK_DAYS, WEEK_DAYS_INVERTED, WIDTH_COLUMNS
from .models import Lesson, ProcessedLesson

logger = logging.getLogger(__name__)


def process_lessons_for_export(
    raw_lessons: list[Lesson], subgroup_name: str, first_day: datetime.date,
) -> list[ProcessedLesson]:
    """
    Process, filter, merge and transform raw lesson data into ProcessedLesson objects.

    Args:
        raw_lessons: List of raw Lesson objects.
        subgroup_name: Subgroup name for filtering.
        first_day: Start date of the schedule.

    Returns:
        Filtered, sorted and processed list of ProcessedLesson objects.
    """
    # 1. Filter by subgroup
    def f_filter(lesson: Lesson) -> bool:
        r: bool = lesson.subgroup == subgroup_name.upper()
        r = r or lesson.subgroup == subgroup_name.lower()
        return r

    filtered_lessons = [lesson for lesson in raw_lessons if f_filter(lesson)]
    processed_lessons: dict[tuple[datetime.date, int, str], ProcessedLesson] = {}
    lesson_counter: dict[tuple[str, str], int] = {}

    for lesson in filtered_lessons:
        week_num = int(lesson.weekNumber)
        day_index = WEEK_DAYS.get(lesson.dayName, 8)

        # Get time data from single source
        if lesson.lessonType.lower() == "лекционного":
            lesson_type_key = "л"
        elif lesson.lessonType.lower() == "семинарского":
            lesson_type_key = "с"
        else:
            lesson_type_key = 'N/A'
            logger.error(f"Found unhandled lesson type: {lesson.lessonType}")

        # --- NEW RELIABLE TIME PROCESSING LOGIC ---
        pair_time = lesson.pairTime.replace(".", ":")
        # 1. Take only start time (part of string before "-")
        start_time_str = pair_time.split("-")[0].strip()
        try:
            t = start_time_str.split(":")
            lesson_start_time = datetime.time(int(t[0]), int(t[1]))
        except (ValueError, IndexError):
            logger.warning(f"Could not parse time: {start_time_str}")
            continue
        
        # 2. Find lesson number by matching START time
        lesson_number = None
        for num, ring in enumerate(RINGS[lesson_type_key]):
            # Compare only with start time of first "half" of the lesson
            if ring[0][0] == lesson_start_time:
                lesson_number = num
                break  # Found lesson, exit loop
        # 3. If number not found, it's second part of lesson or unknown time -> skip
        if lesson_number is None:
            logger.warning(f"Could not find lesson number for time {lesson_start_time} in {lesson_type_key}")
            continue
        # --- END NEW LOGIC ---

        first_monday = first_day - datetime.timedelta(days=first_day.weekday())
        date = first_monday + datetime.timedelta(weeks=week_num - 1, days=day_index)
        lesson_counter[lesson.subjectName, lesson_type_key] = lesson_counter.get((lesson.subjectName, lesson_type_key), 0) + 1
        
        # Create clean ProcessedLesson object
        try:
            time_info = RINGS[lesson_type_key][lesson_number]
            time_slot = [time_info[0][0], time_info[0][1], time_info[1][0], time_info[1][1]]
        except (KeyError, IndexError):
            continue
            
        processed_lesson = ProcessedLesson(
            week=week_num,
            date=date,
            time_slot=time_slot,
            lesson_numbers=str(lesson_number + 1),
            type_=lesson_type_key.upper(),
            subject=lesson.subjectName,
            location=lesson.locationAddress,
            lecturer=lesson.lectorName
        )
        
        key: tuple[datetime.date, int, str] = (
            date,
            lesson_number,
            lesson_type_key,
        )
        if key in processed_lessons.keys():
            logger.warning("WARN! Overwriting lesson")
        processed_lessons[key] = processed_lesson

    # 3. Sort by date and lesson number
    processed_lessons_list: list[ProcessedLesson] = list(processed_lessons.values())
    processed_lessons_list.sort(key=lambda x: (x.date, int(x.lesson_numbers.split('-')[0])))
    logger.info(f"raw_lessons={len(raw_lessons)};filtered_lessons={len(filtered_lessons)};processed_lessons={len(processed_lessons_list)}")
    return processed_lessons_list


def gen_excel_file(schedule_data: List[ProcessedLesson], subgroup_name: str) -> None:
    """
    Generate Excel file from ProcessedLesson objects.

    Args:
        schedule_data: List of ProcessedLesson objects.
        subgroup_name: Subgroup name for file naming.
    """
    if not schedule_data:
        logger.warning("No data for Excel file generation.")
        return

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    if worksheet is None:
        raise IndexError("Could not create worksheet in Excel.")
    worksheet.title = "Schedule"

    # --- Header ---
    worksheet.append([f"Schedule for {subgroup_name}. Generated by Schedule Processor"])
    worksheet.merge_cells("A1:G1")
    worksheet["A1"].alignment = Alignment(horizontal="center")
    worksheet["A1"].font = Font(size=14, name="Roboto")

    # --- Table header ---
    header = ["Week", "Date", "Day", "№", "Time", "Type", "Subject"]
    worksheet.append(header)
    header_font = Font(size=14, name="Roboto")
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    header_align = Alignment(horizontal="center")
    header_border = Border(
        top=Side(style="thin"),
        left=Side(style="thin"),
        right=Side(style="thin"),
        bottom=Side(style="thick"),
    )
    for cell in worksheet[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Styles ---
    thin_border_bottom = Border(bottom=Side(style="thin"))
    thick_border_bottom = Border(bottom=Side(style="thick"))
    lecture_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    seminar_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # --- Fill with data ---
    if not schedule_data:
        return
        
    first_date_in_schedule = schedule_data[0].date
    prev_date: datetime.date | None = None
    prev_week_num: int | None = None
    merge_start_row_date = 3
    merge_start_row_week = 3
    current_row_index = 3

    for lesson in schedule_data:
        # 1. Extract and calculate data
        lesson_date = lesson.date
        week_number = (lesson_date - first_date_in_schedule).days // 7 + 1
        day_name = WEEK_DAYS_INVERTED.get(lesson_date.weekday(), "")
        lesson_number_display = lesson.lesson_numbers
        lesson_type_key = lesson.type_.lower()
        if lesson_type_key == 'с':
            lesson_number_display = '1-2' if lesson_number_display == '1' else '3-4'

        start_time = lesson.time_slot[0]
        end_time = lesson.time_slot[3]
        time_string = f"{start_time.strftime('%H:%M')}-{end_time.strftime('%H:%M')}"

        row_data = [
            week_number,
            lesson_date.strftime('%d.%m.%Y'),
            day_name,
            lesson_number_display,
            time_string,
            lesson.type_,
            lesson.subject
        ]
        worksheet.append(row_data)

        # 2. Merge cells and apply borders
        if prev_week_num is not None and prev_week_num != week_number:
            if merge_start_row_week < current_row_index - 1:
                worksheet.merge_cells(f"A{merge_start_row_week}:A{current_row_index - 1}")
            for col in range(1, 8):
                worksheet.cell(row=current_row_index - 1, column=col).border = thick_border_bottom
            merge_start_row_week = current_row_index

        if prev_date is not None and prev_date != lesson_date:
            if merge_start_row_date < current_row_index - 1:
                worksheet.merge_cells(f"B{merge_start_row_date}:B{current_row_index - 1}")
                worksheet.merge_cells(f"C{merge_start_row_date}:C{current_row_index - 1}")
            if worksheet.cell(row=current_row_index - 1, column=1).border != thick_border_bottom:
                for col in range(1, 8):
                    worksheet.cell(row=current_row_index - 1, column=col).border = thin_border_bottom
            merge_start_row_date = current_row_index

        # 3. Apply styles to current row
        current_cells = worksheet[current_row_index]
        fill_color = lecture_fill if lesson.type_ == "Л" else seminar_fill
        
        for i, cell in enumerate(current_cells):
            # Set font
            if i == 0:
                cell.font = Font(bold=True, size=28, name="Roboto")
            elif i == 2:
                cell.font = Font(bold=True, size=12, name="Roboto")
            else:
                cell.font = Font(name="Roboto", size=12)
            
            # Set alignment
            if i != 6:
                cell.alignment = center_align

            # Set fill
            if i in [3, 4, 5, 6]: # №, Time, Type, Subject
                 cell.fill = fill_color

        prev_date = lesson_date
        prev_week_num = week_number
        current_row_index += 1

    # 4. Final merge
    if merge_start_row_date < current_row_index - 1:
        worksheet.merge_cells(f"B{merge_start_row_date}:B{current_row_index - 1}")
        worksheet.merge_cells(f"C{merge_start_row_date}:C{current_row_index - 1}")
    if merge_start_row_week < current_row_index - 1:
        worksheet.merge_cells(f"A{merge_start_row_week}:A{current_row_index - 1}")

    # --- Column widths ---
    for i, width in enumerate(WIDTH_COLUMNS):
        worksheet.column_dimensions[chr(65 + i)].width = width

    # --- Save ---
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)
    filename = output_dir / f"{subgroup_name}.xlsx"
    logger.info(f"Saving file: {filename}")
    workbook.save(filename)


def gen_ical(schedule_data: List[ProcessedLesson], subgroup_name: str) -> None:
    """
    Generate iCal file from ProcessedLesson objects.

    Args:
        schedule_data: List of ProcessedLesson objects.
        subgroup_name: Subgroup name for file naming.
    """
    if not schedule_data:
        logger.warning("No data for iCal file generation.")
        return
        
    calendar = ics.Calendar()
    moscow_tz = ZoneInfo('Europe/Moscow')

    for lesson in schedule_data:
        # 1. Get start and end time
        start_time = lesson.time_slot[0]
        end_time = lesson.time_slot[3]

        # 2. Create event
        event = ics.Event()
        event.begin = datetime.datetime.combine(lesson.date, start_time, tzinfo=moscow_tz)
        event.end = datetime.datetime.combine(lesson.date, end_time, tzinfo=moscow_tz)
        event.name = f"№{lesson.lesson_numbers} {lesson.type_} {lesson.subject}"
        event.created = datetime.datetime.now(tz=moscow_tz)

        # 3. Add metadata
        event.location = lesson.location or ""
        event.categories = [{"Л": "Lecture", "С": "Seminar"}.get(lesson.type_, "Class"), "SZGMU"]
        
        description_parts = []
        if lesson.lecturer:
            description_parts.append(f"Lecturer: {lesson.lecturer}")
        
        event.description = "\n".join(description_parts)
        calendar.events.add(event)

    # 4. Save file
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)
    filename = output_dir / f"{subgroup_name}.ics"
    
    if not calendar.events:
        logger.warning(f"No events created. File {filename} will not be saved.")
        return
        
    logger.info(f"Saving file: {filename} ({len(calendar.events)} events)")
    with filename.open("wb") as ical_file:
        ical_file.write(calendar.serialize().encode('utf-8'))