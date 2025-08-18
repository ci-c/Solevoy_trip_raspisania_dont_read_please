import openpyxl
from pathlib import Path
from typing import List
import rich
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from config import WIDTH_COLUMNS, RINGS, WEEK_DAYS_INVERTED
from post_lesson import PostLesson
from datetime import date

def gen_excel_file(schedule_data: List[PostLesson], subgroup_name: str) -> None:
    """
    Генерирует файл Excel из списка объектов PostLesson.

    Args:
        schedule_data: Список объектов PostLesson.
        subgroup_name: Имя подгруппы для именования файла.
    """
    if not schedule_data:
        rich.print("[bold red]Нет данных для генерации Excel файла.[/bold red]")
        return

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    if worksheet is None:
        raise IndexError("Не удалось создать рабочий лист в Excel.")
    worksheet.title = "Расписание"

    # --- Заголовок ---
    worksheet.append([f"Расписание для {subgroup_name}. MechNews: https://t.me/mechnews_1k"])
    worksheet.merge_cells("A1:G1")
    worksheet["A1"].alignment = Alignment(horizontal="center")
    worksheet["A1"].font = Font(size=14, name="Roboto")

    # --- Шапка таблицы ---
    header = ["Нед.", "Дата", "День", "№", "Время", "Тип", "Предмет"]
    worksheet.append(header)
    header_font = Font(size=14, name="Roboto")
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    header_align = Alignment(horizontal="center")
    header_border = Border(
        top=Side(style="thin"),
        left=Side(style="thin"),
        right=Side(style="thin"),
        bottom=Side(style="thick"),
    )
    for cell in worksheet[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Стили ---
    thin_border_bottom = Border(bottom=Side(style="thin"))
    thick_border_bottom = Border(bottom=Side(style="thick"))
    lecture_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    seminar_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # --- Заполнение данными ---
    first_date_in_schedule = schedule_data[0].date
    prev_date: date | None = None
    prev_week_num: int | None = None
    merge_start_row_date = 3
    merge_start_row_week = 3
    current_row_index = 3

    for lesson in schedule_data:
        # 1. Извлечение и вычисление данных
        lesson_date = lesson.date
        week_number = (lesson_date - first_date_in_schedule).days // 7 + 1
        day_name = WEEK_DAYS_INVERTED.get(lesson_date.weekday(), "")
        lesson_number_display = lesson.lesson_number + 1
        
        lesson_type_key = lesson.lesson_type.lower()
        try:
            time_info = RINGS[lesson_type_key][lesson.lesson_number]
            start_time = time_info[0][0]
            end_time = time_info[1][1]
            time_string = f"{start_time.strftime('%H:%M')}-{end_time.strftime('%H:%M')}"
        except (KeyError, IndexError):
            time_string = "N/A"

        row_data = [
            week_number,
            lesson_date.strftime('%d.%m.%Y'),
            day_name,
            f"'{lesson_number_display}-{lesson.lesson_seq}", # Добавил ' чтобы Excel не считал это датой
            time_string,
            lesson.lesson_type,
            lesson.subject_name
        ]
        worksheet.append(row_data)

        # 2. Объединение ячеек и применение границ
        if prev_week_num is not None and prev_week_num != week_number:
            if merge_start_row_week < current_row_index - 1:
                worksheet.merge_cells(f"A{merge_start_row_week}:A{current_row_index - 1}")
            for col in range(1, 8):
                worksheet.cell(row=current_row_index - 1, column=col).border = thick_border_bottom
            merge_start_row_week = current_row_index

        if prev_date is not None and prev_date != lesson_date:
            if merge_start_row_date < current_row_index - 1:
                worksheet.merge_cells(f"B{merge_start_row_date}:B{current_row_index - 1}")
                worksheet.merge_cells(f"C{merge_start_row_date}:C{current_row_index - 1}")
            if worksheet.cell(row=current_row_index - 1, column=1).border != thick_border_bottom:
                for col in range(1, 8):
                    worksheet.cell(row=current_row_index - 1, column=col).border = thin_border_bottom
            merge_start_row_date = current_row_index

        # 3. Применение стилей к текущей строке (ИСПРАВЛЕННЫЙ БЛОК)
        current_cells = worksheet[current_row_index]
        fill_color = lecture_fill if lesson.lesson_type == "Л" else seminar_fill
        
        for i, cell in enumerate(current_cells):
            # Установка шрифта
            if i == 0:
                cell.font = Font(bold=True, size=28, name="Roboto")
            elif i == 2:
                cell.font = Font(bold=True, size=12, name="Roboto")
            else:
                cell.font = Font(name="Roboto", size=12)
            
            # Установка выравнивания
            if i != 6:
                cell.alignment = center_align

            # Установка заливки
            if i in [3, 4, 5, 6]: # №, Время, Тип, Предмет
                 cell.fill = fill_color

        prev_date = lesson_date
        prev_week_num = week_number
        current_row_index += 1

    # 4. Финальное объединение
    if merge_start_row_date < current_row_index - 1:
        worksheet.merge_cells(f"B{merge_start_row_date}:B{current_row_index - 1}")
        worksheet.merge_cells(f"C{merge_start_row_date}:C{current_row_index - 1}")
    if merge_start_row_week < current_row_index - 1:
        worksheet.merge_cells(f"A{merge_start_row_week}:A{current_row_index - 1}")

    # --- Ширина колонок ---
    for i, width in enumerate(WIDTH_COLUMNS):
        worksheet.column_dimensions[chr(65 + i)].width = width

    # --- Сохранение ---
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)
    filename = output_dir / f"{subgroup_name}.xlsx"
    rich.print(f"Сохраняю файл: [cyan]{filename}[/cyan]")
    workbook.save(filename)